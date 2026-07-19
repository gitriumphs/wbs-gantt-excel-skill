#!/usr/bin/env python3
"""Generate a Japanese WBS / Gantt-chart Excel workbook from a tasks CSV and a holiday CSV.

Usage (uniform SP length):
    python generate_wbs.py --tasks tasks.csv --holidays holidays.csv --output out.xlsx \
        [--sp-config sp_config.csv] \
        [--chart-start-date 2026-01-19] [--chart-end-date 2026-03-31] \
        [--sp-start-number 26] [--sp-length-workdays 10] [--num-sps N]

Usage (variable-length SPs, each with its own start/end date):
    python generate_wbs.py --tasks tasks.csv --holidays holidays.csv --output out.xlsx \
        --sp-list sp_list.csv [--chart-start-date ...] [--chart-end-date ...]

SP settings (chart_start_date, chart_end_date, sp_start_number, sp_length_workdays, num_sps) can
be supplied via a --sp-config CSV (key,value rows) instead of CLI flags -- see
assets/sp_config_template.csv. Any CLI flag that is explicitly passed overrides the same setting
from --sp-config. chart_end_date and num_sps are two different ways to say "how much calendar to
render" and are mutually exclusive; leave both unset to auto-size the calendar to the tasks.

--sp-list points to a CSV of SP,開始日,終了日 rows (see assets/sp_list_template.csv) that gives
each SP an explicit, independently-sized date range instead of a uniform sp_length_workdays. It is
mutually exclusive with --sp-start-number/--sp-length-workdays/--num-sps (and the matching
sp_config keys); --chart-start-date/--chart-end-date still work on top of it to trim/extend the
rendered calendar.

Only depends on openpyxl (no pandas). See the skill's SKILL.md for the CSV formats.
"""
import argparse
import csv
import datetime as dt
import math
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STEPS = ["コーディング", "実装", "実装後コーダー確認", "試験", "受け入れ試験", "リリース"]

FIRST_DAY_COL = 9  # column I
HOLIDAY_SHEET_LAST_ROW = 500  # generous fixed range for WORKDAY/NETWORKDAYS formulas

HEADER_FILL = PatternFill("solid", fgColor="FF0B5394")
SP_FILL = PatternFill("solid", fgColor="FF990000")
WEEK_FILL = PatternFill("solid", fgColor="FF2F75B5")
DATE_FILL = PatternFill("solid", fgColor="FFBDD7EE")
GROUP_FILL = PatternFill("solid", fgColor="FFCCCCCC")
# Both fgColor and bgColor are set (not just fgColor) because this fill is used inside a
# conditional-formatting dxf: unlike a normal cell style, some Excel builds (notably Excel for
# Mac) render a dxf's "solid" pattern from bgColor rather than fgColor, so a fgColor-only fill
# shows as blank/white there even though it looks correct in openpyxl/Windows Excel.
GANTT_FILL = PatternFill(fill_type="solid", start_color="FF5B9BD5", end_color="FF5B9BD5")

HEADER_FONT = Font(bold=True, size=9, color="FFFFFFFF")
DATE_FONT = Font(bold=True, size=9, color="FF000000")
GROUP_NUM_FONT = Font(bold=True, size=11, color="FF000000")
GROUP_TITLE_FONT = Font(bold=True, size=11, color="FF0000FF")
TASK_FONT = Font(size=10, color="FF434343")
COL_HEADER_FONT = Font(bold=True, size=8, color="FF000000")

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN_LEFT = Border(left=Side(style="thin"))


class InputError(ValueError):
    pass


# ---------------------------------------------------------------------------
# CSV / date helpers
# ---------------------------------------------------------------------------

def read_csv_rows(path):
    last_err = None
    for enc in ("utf-8-sig", "cp932"):
        try:
            with open(path, newline="", encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                fieldnames = reader.fieldnames
            return fieldnames, rows
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise InputError(f"CSVの文字コードを判定できませんでした(utf-8/cp932で読めません): {path} ({last_err})")


def parse_date(s, context=""):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise InputError(f"日付を解釈できません{(' (' + context + ')') if context else ''}: {s!r}")


def is_workday(d, holidays):
    return d.weekday() < 5 and d not in holidays


def next_workday(d, holidays):
    d = d + dt.timedelta(days=1)
    while not is_workday(d, holidays):
        d += dt.timedelta(days=1)
    return d


def roll_to_workday(d, holidays):
    while not is_workday(d, holidays):
        d += dt.timedelta(days=1)
    return d


def add_workdays(start, n, holidays):
    """start must already be a workday. Returns the date n workdays after start (n=0 -> start)."""
    d = start
    for _ in range(n):
        d = next_workday(d, holidays)
    return d


def workday_index(chart_start, date, holidays):
    """0-based position of `date` in the business-day sequence starting at chart_start."""
    if date < chart_start:
        return -1
    idx = 0
    d = chart_start
    while d < date:
        d = next_workday(d, holidays)
        idx += 1
    return idx


# ---------------------------------------------------------------------------
# Loading holidays / tasks
# ---------------------------------------------------------------------------

def load_holidays(path):
    fieldnames, rows = read_csv_rows(path)
    if not fieldnames or "date" not in fieldnames:
        raise InputError(f"祝日CSVに 'date' 列が見つかりません: {path}")
    holidays = {}
    for row in rows:
        raw = (row.get("date") or "").strip()
        if not raw:
            continue
        d = parse_date(raw, context=f"祝日ファイル {path}")
        holidays[d] = (row.get("name") or "").strip()
    return holidays


def load_tasks(path):
    fieldnames, rows = read_csv_rows(path)
    required = ["機能名", "開始日"] + [f"{s}_担当" for s in STEPS] + [f"{s}_日数" for s in STEPS]
    missing = [c for c in required if c not in (fieldnames or [])]
    if missing:
        raise InputError(f"タスクCSVに必要な列がありません: {missing}\n必要な列: {required}")

    features = []
    for i, row in enumerate(rows, start=2):
        name = (row.get("機能名") or "").strip()
        if not name:
            continue
        raw_start = parse_date(row.get("開始日"), context=f"行{i} ({name}) の開始日")
        steps = []
        for step in STEPS:
            assignee = (row.get(f"{step}_担当") or "").strip()
            if not assignee:
                continue
            dur_raw = (row.get(f"{step}_日数") or "").strip()
            if not dur_raw:
                raise InputError(f"行{i} ({name}): {step}_日数 が未入力です(担当が入っているのに日数が空)")
            try:
                dur = int(dur_raw)
            except ValueError:
                raise InputError(f"行{i} ({name}): {step}_日数 は整数で指定してください: {dur_raw!r}")
            if dur <= 0:
                raise InputError(f"行{i} ({name}): {step}_日数 は正の整数にしてください: {dur}")
            steps.append({"name": step, "assignee": assignee, "duration": dur})
        if not steps:
            raise InputError(f"行{i} ({name}): 少なくとも1つのステップに担当者を指定してください")
        features.append({"name": name, "raw_start": raw_start, "steps": steps})

    if not features:
        raise InputError(f"タスクCSVに有効な行がありません: {path}")
    return features


SP_CONFIG_KEYS = {"chart_start_date", "chart_end_date", "sp_start_number", "sp_length_workdays", "num_sps"}


def load_sp_config(path):
    """Load SP settings from a key,value CSV. Returns (config dict of raw strings, unknown keys)."""
    fieldnames, rows = read_csv_rows(path)
    if not fieldnames or "key" not in fieldnames or "value" not in fieldnames:
        raise InputError(f"SP設定CSVは 'key,value' の形式にしてください: {path}")
    config = {}
    unknown = []
    for row in rows:
        key = (row.get("key") or "").strip()
        value = (row.get("value") or "").strip()
        if not key or not value:
            continue
        if key not in SP_CONFIG_KEYS:
            unknown.append(key)
            continue
        config[key] = value
    return config, unknown


def parse_int_setting(value, field_name):
    try:
        return int(value)
    except (TypeError, ValueError):
        raise InputError(f"{field_name} は整数で指定してください: {value!r}")


def load_sp_list(path):
    """Load an explicit per-SP date-range list from a SP,開始日,終了日 CSV.

    Each row gives one SP's own start/end date, so SPs can have different lengths. Returns a list
    of {"sp": int, "start": date, "end": date} sorted by start date.
    """
    fieldnames, rows = read_csv_rows(path)
    required = ["SP", "開始日", "終了日"]
    missing = [c for c in required if c not in (fieldnames or [])]
    if missing:
        raise InputError(f"SP一覧CSVに必要な列がありません: {missing}\n必要な列: {required}")

    entries = []
    for i, row in enumerate(rows, start=2):
        sp_raw = (row.get("SP") or "").strip()
        if not sp_raw:
            continue
        sp_number = parse_int_setting(sp_raw, f"SP一覧 行{i} の SP番号")
        start = parse_date(row.get("開始日"), context=f"SP一覧 行{i} (SP{sp_number}) の開始日")
        end = parse_date(row.get("終了日"), context=f"SP一覧 行{i} (SP{sp_number}) の終了日")
        if end < start:
            raise InputError(f"SP一覧 行{i} (SP{sp_number}): 終了日({end})が開始日({start})より前です")
        entries.append({"sp": sp_number, "start": start, "end": end})

    if not entries:
        raise InputError(f"SP一覧CSVに有効な行がありません: {path}")

    entries.sort(key=lambda e: e["start"])
    for a, b in zip(entries, entries[1:]):
        if a["end"] >= b["start"]:
            raise InputError(
                f"SP一覧: SP{a['sp']}({a['start']}〜{a['end']}) と "
                f"SP{b['sp']}({b['start']}〜{b['end']}) の期間が重複しています"
            )
    return entries


def sp_list_gap_warnings(entries, holidays):
    """Business days between consecutive SPs that no SP row covers (informational only)."""
    warnings = []
    for a, b in zip(entries, entries[1:]):
        d = next_workday(a["end"], holidays)
        if d < b["start"]:
            warnings.append(
                f"SP{a['sp']}の終了({a['end']})とSP{b['sp']}の開始({b['start']})の間に、"
                f"どのSPにも属さない営業日があります"
            )
    return warnings


def sp_label_for_date(d, entries):
    for e in entries:
        if e["start"] <= d <= e["end"]:
            return e["sp"]
    return None


def schedule_features(features, holidays, adjustments):
    for feat in features:
        start = roll_to_workday(feat["raw_start"], holidays)
        if start != feat["raw_start"]:
            adjustments.append(f"「{feat['name']}」: 開始日 {feat['raw_start']} は非稼働日のため {start} に調整しました")
        feat["start"] = start
        cur_start = start
        for step in feat["steps"]:
            step["start"] = cur_start
            step["end"] = add_workdays(cur_start, step["duration"] - 1, holidays)
            cur_start = next_workday(step["end"], holidays)
    return features


# ---------------------------------------------------------------------------
# Calendar / SP grid
# ---------------------------------------------------------------------------

def build_calendar(chart_start, holidays, sp_length, num_sps_override, chart_end, features):
    if num_sps_override and chart_end:
        raise InputError("終了日(chart_end_date)とSP数(num_sps)は同時に指定できません。どちらか一方にしてください。")

    if num_sps_override:
        num_sps = num_sps_override
    elif chart_end:
        if chart_end < chart_start:
            raise InputError(f"終了日(chart_end_date={chart_end})が開始日(chart_start_date={chart_start})より前です。")
        needed = workday_index(chart_start, chart_end, holidays) + 1
        num_sps = math.ceil(needed / sp_length)
    else:
        all_ends = [s["end"] for f in features for s in f["steps"]]
        max_end = max(all_ends) if all_ends else chart_start
        needed = workday_index(chart_start, max_end, holidays) + 1
        if needed < 1:
            needed = 1
        num_sps = math.ceil(needed / sp_length) + 1  # +1 SP of buffer, auto-sizing mode only

    total_workdays = num_sps * sp_length
    calendar = [chart_start]
    d = chart_start
    for _ in range(total_workdays - 1):
        d = next_workday(d, holidays)
        calendar.append(d)
    return calendar, num_sps


def build_calendar_explicit(chart_start, chart_end, holidays):
    """Business-day calendar spanning chart_start..chart_end inclusive (for --sp-list mode).

    Chains next_workday() and stops once it would pass chart_end, rather than pre-computing a
    day count via workday_index() -- that count-based approach overshoots by one whenever
    chart_end itself falls on a weekend/holiday (it rounds up to the next workday on or after
    chart_end instead of stopping at the last workday on or before it).
    """
    if chart_end < chart_start:
        raise InputError(f"終了日(chart_end_date={chart_end})が開始日(chart_start_date={chart_start})より前です。")
    calendar = [chart_start]
    d = chart_start
    while True:
        nxt = next_workday(d, holidays)
        if nxt > chart_end:
            break
        calendar.append(nxt)
        d = nxt
    return calendar


def week_of_month(d):
    return (d.day - 1) // 7 + 1


def merge_band(ws, row, calendar, key_func, label_func, fill, font, align=CENTER):
    col = FIRST_DAY_COL
    i = 0
    n = len(calendar)
    while i < n:
        key = key_func(calendar[i])
        j = i
        while j + 1 < n and key_func(calendar[j + 1]) == key:
            j += 1
        start_col = FIRST_DAY_COL + i
        end_col = FIRST_DAY_COL + j
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = THIN_LEFT
        if end_col > start_col:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.cell(row=row, column=start_col, value=label_func(calendar[i]))
        i = j + 1


def merge_sp_band(ws, row, sp_labels, fill, font):
    """Render the SP band from a precomputed per-day label list (works for both uniform SP
    length and explicit --sp-list mode -- consecutive equal labels get merged into one cell)."""
    n = len(sp_labels)
    i = 0
    while i < n:
        label = sp_labels[i]
        j = i
        while j + 1 < n and sp_labels[j + 1] == label:
            j += 1
        start_col = FIRST_DAY_COL + i
        end_col = FIRST_DAY_COL + j
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = CENTER
            cell.border = THIN_LEFT
        if end_col > start_col:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.cell(row=row, column=start_col, value=(f"SP{label}" if label is not None else ""))
        i = j + 1


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------

def build_workbook(features, holidays, chart_start, sp_start_number, sp_length, num_sps_override,
                    chart_end=None, sp_entries=None):
    adjustments = []
    schedule_features(features, holidays, adjustments)

    for feat in features:
        for step in feat["steps"]:
            if workday_index(chart_start, step["start"], holidays) < 0:
                raise InputError(
                    f"「{feat['name']}」の「{step['name']}」の開始日({step['start']})が"
                    f"チャート開始日({chart_start})より前です。chart_start_date を早めてください。"
                )

    sp_warnings = []
    if sp_entries is not None:
        if chart_end is None:
            chart_end = max(e["end"] for e in sp_entries)
        calendar = build_calendar_explicit(chart_start, chart_end, holidays)
        num_sps = len(sp_entries)
        sp_labels = [sp_label_for_date(d, sp_entries) for d in calendar]
        sp_warnings.extend(sp_list_gap_warnings(sp_entries, holidays))
        if any(lbl is None for lbl in sp_labels):
            sp_warnings.append("表示範囲内にどのSPにも属さない営業日があります(SP列は空欄になります)")
        sp_numbers = sorted(e["sp"] for e in sp_entries)
        sp_range_label = f"SP{sp_numbers[0]}-SP{sp_numbers[-1]}"

        def sp_number_for(d):
            return sp_label_for_date(d, sp_entries)
    else:
        calendar, num_sps = build_calendar(chart_start, holidays, sp_length, num_sps_override, chart_end, features)
        sp_labels = [sp_start_number + i // sp_length for i in range(len(calendar))]
        sp_range_label = f"SP{sp_start_number}-SP{sp_start_number + num_sps - 1}"

        def sp_number_for(d):
            return sp_start_number + workday_index(chart_start, d, holidays) // sp_length

    overflow = []
    for feat in features:
        for step in feat["steps"]:
            if step["end"] > calendar[-1]:
                overflow.append(
                    f"「{feat['name']}」の「{step['name']}」の終了日({step['end']})が"
                    f"表示範囲の終了({calendar[-1]})を超えています"
                )
    total_workdays = len(calendar)
    last_col = FIRST_DAY_COL + total_workdays - 1
    last_col_letter = get_column_letter(last_col)
    first_col_letter = get_column_letter(FIRST_DAY_COL)

    wb = Workbook()
    ws = wb.active
    ws.title = "ガントチャート"

    # --- fixed left-hand headers (A1:H5) ---
    headers = ["項番", "タスクのタイトル", "担当", "SP", "開始", "終了", "期間", "タスク\n完了率"]
    for idx, label in enumerate(headers, start=1):
        for r in range(1, 6):
            cell = ws.cell(row=r, column=idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.merge_cells(start_row=1, start_column=idx, end_row=5, end_column=idx)
        top = ws.cell(row=1, column=idx, value=label)
        top.alignment = CENTER_WRAP

    # --- month / SP / week / date / weekday bands ---
    merge_band(ws, 1, calendar, key_func=lambda d: (d.year, d.month),
               label_func=lambda d: f"{d.month}月", fill=HEADER_FILL, font=HEADER_FONT)
    merge_sp_band(ws, 2, sp_labels, fill=SP_FILL, font=HEADER_FONT)
    merge_band(ws, 3, calendar, key_func=lambda d: (d.year, d.month, week_of_month(d)),
               label_func=lambda d: f"第 {week_of_month(d)} 週", fill=WEEK_FILL, font=HEADER_FONT)

    for i, d in enumerate(calendar):
        col = FIRST_DAY_COL + i
        letter = get_column_letter(col)
        cell = ws.cell(row=4, column=col)
        if i == 0:
            cell.value = d
        else:
            prev_letter = get_column_letter(col - 1)
            cell.value = f"=WORKDAY({prev_letter}4,1,'設定'!$A$2:$A${HOLIDAY_SHEET_LAST_ROW})"
        cell.number_format = "d"
        cell.fill = DATE_FILL
        cell.font = DATE_FONT
        cell.alignment = CENTER
        cell.border = THIN_LEFT

        wcell = ws.cell(row=5, column=col, value=f'=TEXT({letter}4,"ddd")')
        wcell.fill = DATE_FILL
        wcell.font = DATE_FONT
        wcell.alignment = CENTER
        wcell.border = THIN_LEFT

    # --- group / task rows ---
    row = 6
    for gi, feat in enumerate(features, start=1):
        ws.cell(row=row, column=1, value=gi).font = GROUP_NUM_FONT
        ws.cell(row=row, column=2, value=feat["name"]).font = GROUP_TITLE_FONT
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = GROUP_FILL
        ws.cell(row=row, column=1).alignment = LEFT
        ws.row_dimensions[row].height = 21
        row += 1

        group_start_row = row
        for si, step in enumerate(feat["steps"], start=1):
            sp_number = sp_number_for(step["start"])
            values = {
                1: si,
                2: step["name"],
                3: step["assignee"],
                4: sp_number if sp_number is not None else "",
            }
            for col, val in values.items():
                c = ws.cell(row=row, column=col, value=val)
                c.font = TASK_FONT
                c.alignment = CENTER if col in (1, 4) else LEFT

            ecell = ws.cell(row=row, column=5, value=step["start"])
            ecell.number_format = 'm"/"d'
            ecell.font = TASK_FONT
            ecell.alignment = LEFT

            fcell = ws.cell(row=row, column=6, value=step["end"])
            fcell.number_format = 'm"/"d'
            fcell.font = TASK_FONT
            fcell.alignment = LEFT

            gcell = ws.cell(row=row, column=7,
                             value=f"=NETWORKDAYS(E{row},F{row},'設定'!$A$2:$A${HOLIDAY_SHEET_LAST_ROW})")
            gcell.font = TASK_FONT
            gcell.alignment = CENTER

            hcell = ws.cell(row=row, column=8, value=0)
            hcell.number_format = "0%"
            hcell.font = TASK_FONT
            hcell.alignment = CENTER

            row += 1
        group_end_row = row - 1

        formula = f"AND({first_col_letter}$4>=$E{group_start_row},{first_col_letter}$4<=$F{group_start_row})"
        rng = f"{first_col_letter}{group_start_row}:{last_col_letter}{group_end_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[formula], fill=GANTT_FILL))

    last_row = row - 1
    if last_row >= 6:
        ws.conditional_formatting.add(
            f"H6:H{last_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFF8696B",
                            mid_type="num", mid_value=0.5, mid_color="FFFFEB84",
                            end_type="num", end_value=1, end_color="FF63BE7B"),
        )

    # --- column widths / row heights / freeze panes ---
    widths = {"A": 6.13, "B": 27.75, "C": 10.0, "D": 4.38, "E": 5.0, "F": 5.0, "G": 5.13, "H": 6.38}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    for col in range(FIRST_DAY_COL, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 3.0
    for r in range(1, 6):
        ws.row_dimensions[r].height = 17.25
    ws.freeze_panes = "I6"

    # --- 設定 sheet (holidays) ---
    ws2 = wb.create_sheet("設定")
    for i, (d, name) in enumerate(sorted(holidays.items()), start=2):
        ws2.cell(row=i, column=1, value=dt.datetime(d.year, d.month, d.day)).number_format = "yyyy/m/d"
        ws2.cell(row=i, column=2, value=name)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16

    summary = {
        "features": len(features),
        "tasks": sum(len(f["steps"]) for f in features),
        "chart_start": chart_start,
        "chart_end": calendar[-1],
        "num_sps": num_sps,
        "sp_range": sp_range_label,
        "adjustments": adjustments,
        "overflow": overflow,
        "sp_warnings": sp_warnings,
    }
    return wb, summary


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Generate a Japanese WBS/Gantt xlsx from a tasks CSV and holiday CSV.")
    p.add_argument("--tasks", required=True, help="Wide-format tasks CSV path")
    p.add_argument("--holidays", required=True, help="Holiday CSV path (columns: date,name)")
    p.add_argument("--output", required=True, help="Output .xlsx path")
    p.add_argument("--sp-config", default=None,
                    help="SP settings CSV (key,value rows: chart_start_date/chart_end_date/"
                         "sp_start_number/sp_length_workdays/num_sps). CLI flags below override "
                         "matching keys from this file.")
    p.add_argument("--sp-start-number", type=int, default=None,
                    help="SP number of the first sprint (default: 26, or --sp-config)")
    p.add_argument("--sp-length-workdays", type=int, default=None,
                    help="Business days per SP (default: 10, or --sp-config)")
    p.add_argument("--sp-list", default=None,
                    help="CSV of SP,開始日,終了日 rows giving each SP its own explicit date range "
                         "so SPs can vary in length. Mutually exclusive with --sp-start-number/"
                         "--sp-length-workdays/--num-sps (and the matching --sp-config keys).")
    p.add_argument("--chart-start-date", default=None,
                    help="First day shown on the chart, YYYY-MM-DD (default: --sp-config, or the "
                         "earliest 開始日 in the tasks CSV)")
    p.add_argument("--chart-end-date", default=None,
                    help="Last day the calendar must cover, YYYY-MM-DD. Mutually exclusive with "
                         "--num-sps (default: --sp-config, or unset)")
    p.add_argument("--num-sps", type=int, default=None,
                    help="Number of SPs to render. Mutually exclusive with --chart-end-date "
                         "(default: --sp-config, or auto-computed to cover all tasks + 1 buffer SP)")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    try:
        config, unknown_keys = ({}, [])
        if args.sp_config:
            config, unknown_keys = load_sp_config(args.sp_config)

        sp_entries = None
        if args.sp_list:
            conflicting = []
            if args.sp_start_number is not None:
                conflicting.append("--sp-start-number")
            if args.sp_length_workdays is not None:
                conflicting.append("--sp-length-workdays")
            if args.num_sps is not None:
                conflicting.append("--num-sps")
            for key in ("sp_start_number", "sp_length_workdays", "num_sps"):
                if key in config:
                    conflicting.append(f"sp_config:{key}")
            if conflicting:
                raise InputError(
                    "--sp-list はSPごとに開始日/終了日を指定するモードのため、均一SP長の設定と"
                    f"同時には指定できません: {', '.join(conflicting)}"
                )
            sp_entries = load_sp_list(args.sp_list)

        sp_start_number = (
            args.sp_start_number if args.sp_start_number is not None
            else parse_int_setting(config["sp_start_number"], "sp_start_number") if "sp_start_number" in config
            else 26
        )
        sp_length_workdays = (
            args.sp_length_workdays if args.sp_length_workdays is not None
            else parse_int_setting(config["sp_length_workdays"], "sp_length_workdays") if "sp_length_workdays" in config
            else 10
        )
        if sp_length_workdays <= 0:
            raise InputError("sp_length_workdays は正の整数にしてください")

        num_sps_override = (
            args.num_sps if args.num_sps is not None
            else parse_int_setting(config["num_sps"], "num_sps") if "num_sps" in config
            else None
        )

        chart_start_raw = args.chart_start_date or config.get("chart_start_date")
        chart_end_raw = args.chart_end_date or config.get("chart_end_date")
        chart_end = parse_date(chart_end_raw, context="終了日(chart_end_date)") if chart_end_raw else None

        holidays = load_holidays(args.holidays)
        features = load_tasks(args.tasks)

        chart_start_auto = False
        if chart_start_raw:
            chart_start = parse_date(chart_start_raw, context="開始日(chart_start_date)")
        elif sp_entries is not None:
            chart_start = min(e["start"] for e in sp_entries)
            chart_start_auto = True
        else:
            chart_start = min(f["raw_start"] for f in features)
            chart_start_auto = True

        wb, summary = build_workbook(
            features, holidays, chart_start,
            sp_start_number, sp_length_workdays, num_sps_override, chart_end,
            sp_entries=sp_entries,
        )

        out_dir = os.path.dirname(os.path.abspath(args.output))
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        wb.save(args.output)
    except InputError as e:
        print(f"エラー: {e}", file=sys.stderr)
        return 1

    print(f"作成しました: {args.output}")
    print(f"機能数: {summary['features']} / タスク数: {summary['tasks']}")
    if chart_start_auto:
        source = "SP一覧の最早開始日" if sp_entries is not None else "タスクの最早開始日"
        print(f"(開始日が未指定のため、{source} {summary['chart_start']} を使用しました)")
    print(f"期間: {summary['chart_start']} 〜 {summary['chart_end']} ({summary['sp_range']}, 全{summary['num_sps']}SP)")
    if unknown_keys:
        print(f"注意: --sp-config の不明なキーを無視しました: {unknown_keys}")
    if summary["adjustments"]:
        print("開始日の調整:")
        for a in summary["adjustments"]:
            print(f"  - {a}")
    if summary["overflow"]:
        print("表示範囲を超えるタスク(chart_end_date/num_spsを広げてください):")
        for o in summary["overflow"]:
            print(f"  - {o}")
    if summary["sp_warnings"]:
        print("SPに関する注意:")
        for w in summary["sp_warnings"]:
            print(f"  - {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
